# Se ejecuta como $python GranoMeter.py
import Jetson.GPIO as GPIO
import time
import os
from Tkinter import *
import ttk
import multiprocessing 
from skimage.feature import peak_local_max
from skimage.morphology import watershed
from scipy import ndimage
import numpy as np
import cv2
import math
import datetime
from functools import partial
import serial
from openpyxl import Workbook


#Variable globales flags
cerrartodo=0
semilla=0

#Version del software
__version__ = '3.2.1'

## Definicion de pines de salida
output_luz = 18  # BOARD pin 12, BCM pin 18
#output_motor1 = 20  # BOARD pin 38, BCM pin 20
#output_motor2 = 21  # BOARD pin 40, BCM pin 21

## Definicion de Semillas
SOJA = 0
COLZA = 1
VICIA = 2
MAIZ = 3
TRIGO = 4
TRIGOSARRACENO = 5
POROTOB = 6
POROTON = 7
POROTOR = 8
GARBANZO = 9
CHIA = 10
AVENA =11
CEBADA = 12
GIRASOL =13
QUINOA = 14
SESAMO = 18
#AREA = 15
#AREAC = 16

# Pin Setup:
GPIO.setmode(GPIO.BCM)
# Seteo de pines en 0 e Inicializacion
GPIO.setup(output_luz, GPIO.OUT, initial=GPIO.LOW)
#GPIO.setup(output_motor1, GPIO.OUT, initial=GPIO.LOW)
#GPIO.setup(output_motor2, GPIO.OUT, initial=GPIO.LOW)


#Proceso Watersheed 
def watersheed(numbers, q, q2,  pipe_ws):
    #tomo los valores del pipe
    output_p_ws, input_p_ws=pipe_ws
    output_p_ws.close()
    global ventana
    global cerrartodo
    medicion=0
    
    #Excel con Tablas
    vhora=datetime.datetime.now()
    wb2=Workbook()
    ruta2=("/home/proyecto/Desktop/Mediciones/Granometer {:02d}:{:02d}:{:02d} {:02d}-{:02d}-{:04d}.xlsx".format(vhora.hour, vhora.minute, vhora.second, vhora.day, vhora.month, vhora.year))
    hoja2=wb2.active
    hoja2.title="Mediciones"
    fila2=1
    col_area2=1
    hoja2.cell(column=col_area2, row=fila2, value="GRANOMETER")
    fila2=3
    column = str(chr(64 + col_area2))
    hoja2.column_dimensions[column].width = 11
    hoja2.cell(column=col_area2, row=fila2, value="FECHA")
    col_area2+=1
    column = str(chr(64 + col_area2))
    hoja2.column_dimensions[column].width = 9
    hoja2.cell(column=col_area2, row=fila2, value="HORA")
    col_area2+=1
    column = str(chr(64 + col_area2))
    hoja2.column_dimensions[column].width = 18
    hoja2.cell(column=col_area2, row=fila2, value="ESPECIE")
    col_area2+=1
    column = str(chr(64 + col_area2))
    hoja2.column_dimensions[column].width = 9
    hoja2.cell(column=col_area2, row=fila2, value="MUESTRA")
    col_area2+=1
    column = str(chr(64 + col_area2))
    hoja2.column_dimensions[column].width = 11
    hoja2.cell(column=col_area2, row=fila2, value="REPETICION")
    col_area2+=1
    column = str(chr(64 + col_area2))
    hoja2.column_dimensions[column].width = 20
    hoja2.cell(column=col_area2, row=fila2, value="PESO PROMEDIO(mg)")
    col_area2+=1
    column = str(chr(64 + col_area2))
    hoja2.column_dimensions[column].width = 10
    hoja2.cell(column=col_area2, row=fila2, value="CANTIDAD")
    col_area2+=1
    column = str(chr(64 + col_area2))
    hoja2.column_dimensions[column].width = 21
    hoja2.cell(column=col_area2, row=fila2, value="AREA PROMEDIO(mm2)")
    wb2.save(filename=ruta2)
    
    while (1):
        medicion=q.get()
        if (medicion==5): 
            return 0
        if medicion:
            #recibo el peso
            peso=q.get() 
            muestra=q.get()
            repeticion=q.get() 
            #Pongo barra en 0
            input_p_ws.send('0') 
            # Apago Vibracion y luz
            #GPIO.output(output_motor1, GPIO.LOW)
            #GPIO.output(output_motor2, GPIO.LOW)
            GPIO.output(output_luz, GPIO.LOW)
            # Verifico que semilla fue seleccionada y se ejecuta la correspondiente
            semilla_selec=q.get()#Espero dato de semilla seleccionada
            nombresemilla="semilla"
            GPIO.output(output_luz, GPIO.HIGH) #Enciendo Luz para sacar la foto
            # Saco la foto
            os.system("gphoto2 --auto-detect")
            os.system("gvfs-mount -l")
            os.system("gvfs-mount -s gphoto2")
            os.system("gphoto2 --auto-detect")
            os.system("gphoto2 --set-config capturetarget=0")
            os.system("gphoto2 --capture-image-and-download --filename '/home/proyecto/GranoMeter/Capturas/myphoto.jpg' --force-overwrite")
            GPIO.output(output_luz, GPIO.LOW)
            os.system("gphoto2 --set-config capturetarget=1")   
            
            v_Blur=0 #Flag para seleccion de blur
            
            #Asignamos parametros a la semilla seleccionada
            if semilla_selec==COLZA:
                v_min_dist=10 
                v_BlurxKernel=29
                v_BluryKernel=29
                v_CountorDistance=15
                v_factor_t= 1.25
                nombresemilla="COLZA"
            elif semilla_selec==QUINOA:
                v_min_dist=10 
                v_BlurxKernel=29
                v_BluryKernel=29
                v_CountorDistance=10
                v_factor_t= 1.25
                nombresemilla="QUINOA" 
            elif semilla_selec==VICIA:
                v_min_dist=10 
                v_BlurxKernel=29
                v_BluryKernel=29
                v_CountorDistance=15
                v_factor_t= 1.25
                nombresemilla="VICIA"    
            elif semilla_selec==TRIGO:
                v_min_dist=15  
                v_BlurxKernel=45
                v_BluryKernel=45 
                v_CountorDistance=30
                v_factor_t= 1.08
                nombresemilla="TRIGO"
            elif semilla_selec==CEBADA:
                v_min_dist=15  
                v_BlurxKernel=45
                v_BluryKernel=45 
                v_CountorDistance=25
                v_factor_t= 1.15
                nombresemilla="CEBADA" 
            elif semilla_selec==AVENA:
                v_min_dist=32  
                v_BlurxKernel=69
                v_BluryKernel=69 
                v_CountorDistance=25
                v_factor_t= 1.1
                nombresemilla="AVENA"         
            elif semilla_selec==GIRASOL:
                v_min_dist=30
                v_BlurxKernel=99
                v_BluryKernel=99
                v_CountorDistance=50
                v_factor_t= 1.15
                nombresemilla="GIRASOL"
            elif semilla_selec==POROTOB:
                v_min_dist=80
                v_BlurxKernel=89
                v_BluryKernel=89
                v_CountorDistance=140
                v_factor_t= 1.08
                nombresemilla="POROTO BLANCO"
            elif semilla_selec==TRIGOSARRACENO:
                v_min_dist=15  
                v_BlurxKernel=69
                v_BluryKernel=69 
                v_CountorDistance=35 
                v_factor_t= 1.08
                nombresemilla="TRIGO SARRACENO"
            elif semilla_selec==MAIZ:
                v_min_dist=20  
                v_BlurxKernel=89
                v_BluryKernel=89 
                v_CountorDistance=45
                v_factor_t= 1.2
                nombresemilla="MAIZ"
            elif semilla_selec==SOJA:
                v_min_dist=20  
                v_BlurxKernel=89
                v_BluryKernel=89 
                v_CountorDistance=35
                v_factor_t= 1.08
                nombresemilla="SOJA"
            elif semilla_selec==GARBANZO:
                v_min_dist=20  
                v_BlurxKernel=89
                v_BluryKernel=89 
                v_CountorDistance=35
                v_factor_t= 1.08
                nombresemilla="GARBANZO"    
            elif semilla_selec==POROTON:
                v_min_dist=20  
                v_BlurxKernel=89
                v_BluryKernel=89 
                v_CountorDistance=50
                v_factor_t= 1.3
                nombresemilla="POROTO NEGRO"
            elif semilla_selec==POROTOR:
                v_min_dist=30  
                v_BlurxKernel=129
                v_BluryKernel=129
                v_CountorDistance=70
                v_factor_t= 1.3
                nombresemilla="POROTO ROJO"
            elif semilla_selec==CHIA:
                v_min_dist=10  
                v_BlurxKernel=17
                v_BluryKernel=17 
                v_CountorDistance=2
                v_factor_t= 1.2
                nombresemilla="CHIA"
            #elif semilla_selec==AREA:
            #  v_Blur=1
            #  v_min_dist=120
            #  v_ThreshImage=160
            #  v_BlurxImage=59
            #  v_BluryImage=59
            #  v_BlurxKernel=89
            #  v_BluryKernel=89  
            #  v_CountorDistance=0
            #  v_factor_t= 1.3
            
            #elif semilla_selec==AREAC:
            #   v_Blur=1
            #   v_min_dist=40
            #   v_ThreshImage=160
            #   v_BlurxImage=5
            #   v_BluryImage=5
            #   v_BlurxKernel=89
            #   v_BluryKernel=89  
            #   v_CountorDistance=0
            #   v_factor_t= 1.08
        
            input_p_ws.send('10') #Proceso 10%
            
            # Leo la imagen y la doy vuelta
            image = cv2.imread("/home/proyecto/GranoMeter/Capturas/myphoto.jpg")
            (h,w)=image.shape[:2]
            center=(w/2,h/2)
            M2=cv2.getRotationMatrix2D(center,180,1)
            image=cv2.warpAffine(image,M2,(w,h))
            imagesal=image
            image=cv2.bitwise_not(image)
            
            input_p_ws.send('20') #Procesando 20%
            
            #Convierto la imagen a escala de grises 
            #verifico si tiene binarizacion manual y Blur o Binarizacion OTSU
            gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
            #cv2.imwrite('/home/proyecto/GranoMeter/imagensalida/salidagray.png',gray)
            if v_Blur:
                th1,thresh = cv2.threshold(gray, v_ThreshImage, 255,cv2.THRESH_BINARY)
                thresh_G = cv2.GaussianBlur(thresh,(v_BlurxImage,v_BluryImage),0) 	
            else:
                thresh = cv2.threshold(gray, 0, 255,cv2.THRESH_BINARY | cv2.THRESH_OTSU)[1]
                thresh_G = thresh
            cv2.imwrite('/home/proyecto/GranoMeter/imagensalida/salidatresh.png',thresh)

            input_p_ws.send('40') #Procesando 40%

            # Calculo la discantia euclediana exacta de cada pixel binario
            # al pixel cero mas cercano, luego encuentro los picos en
            # este mapa de distancia
            D = ndimage.distance_transform_edt(thresh_G) 
            # Aplico Blur a dicha a la transformada de distancia
            #cv2.imwrite('/home/proyecto/GranoMeter/imagensalida/salidatransdist.png',D)
            D = cv2.GaussianBlur(D,(v_BlurxKernel,v_BluryKernel),0) #el kernel debe ser impar
            D = D.astype('uint8')
            #cv2.imwrite('/home/proyecto/GranoMeter/imagensalida/salidatransdistblur.png',D)
            #Ajustamos el valor de min distancia correspondiente a la semilla
            localMax = peak_local_max(D, indices=False, min_distance=v_min_dist,
                labels=thresh)
       
            input_p_ws.send('60') #Procesando 60%
            
            # Realizo un analisis de componentes conectados en los picos locales
            # Usando 8-conectividad, aplico algoritmo watersheed
            markers = ndimage.label(localMax, structure=np.ones((3, 3)))[0]
            labels = watershed(-D, markers, mask=thresh)
            # tomo el valor de etiquetas devuelta por el watersheed

            input_p_ws.send('80') #Procesando 80%

            # Variables algoritmo de medicion
            semillas=0
            cant=0
            areat=0
            area_media=0
            p1=[0, 0]
            p2=[0, 0]
            areas=[]
            areagraf=[]
            cX=[]
            cY=[]
            # Verifico las etiquetas devueltas por watersheed
            for label in np.unique(labels):
                # Si la etiqueta es 0 es por que estamos en el fondo
                # Por lo que lo ignoramos
                if label == 0:
                    continue
                #De no ser 0 corresponde a un contorno valido lo cuento como semilla
                semillas=semillas+1
                # Asigno memoria para la region de la etiqueta y dibujo en la mascara
                mask = np.zeros(gray.shape, dtype="uint8")
                mask[labels == label] = 255
                
                # Detecto los contornos en la mascara y guardo el mas grande
                contours = cv2.findContours(mask.copy(), cv2.RETR_EXTERNAL,
                    cv2.CHAIN_APPROX_SIMPLE)[-2]
                imagesal=cv2.drawContours(imagesal,contours,-1,(0,255,0),2)
                markers=cv2.drawContours(markers,contours,-1,(255,255,255),2)
                c = max(contours, key=cv2.contourArea)
                M = cv2.moments(c)
                if M["m00"]!=0:
                    cX.append(int(M["m10"] / M["m00"])) 
                    cY.append(int(M["m01"] / M["m00"])) 
                else:
                    cX.append(int(M["m00"])) 
                    cY.append(int(M["m00"])) 	
                
                #Inicializo los vectores de las Areas
                for cnt in contours:
                    areas.append(cv2.contourArea(cnt))
                    areagraf.append(cv2.contourArea(cnt))
            
            #Realizo el analisis de los contornos si corresponde a una o mas semilla
            #Le pongo el numero a cada una y voy sumando las areas para el calculo
            # del area promedio, guardo el valor de cada area para luego enviar al xls        
            for i in range(0, semillas):  
                areat=areat+areas[i]
                area_media=(areat/semillas)*v_factor_t #lo uso como margen para no contar de mas semillas grandes       
            areat=(areat*0.0065036400258046845)
            for i in range(0, semillas):
                semilla_check=1
                areabuff=areas[i]*0.0065036400258046845
                areabuff=float("{:.2f}".format(areabuff))
                areagraf[i]=areabuff
                if i>1:
                    p2 = [cX[i], cY[i]]
                    for j in range(1, (i-1)):
                        p1 = [cX[i-j], cY[i-j]]
                        distance= math.sqrt( ((p1[0]-p2[0])**2)+((p1[1]-p2[1])**2) )
                        if distance<v_CountorDistance: 
                            semilla_check=0 #si la distancia entre centros es menor del umbral la descarta como semilla al segmento actual.
                            continue
                if areas[i] >(area_media/5) and areas[i] <(area_media*4)  and semilla_check==1:
                    if areas[i]>area_media:
                        cant_aux=round((areas[i]/area_media))
                        if cant_aux==1:
                            cv2.putText(imagesal, "#{}".format(int(cant)), (int(cX[i]) - 10, int(cY[i])),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.4, (0, 0, 255), 1)
                        elif cant_aux==2:
                            cv2.putText(imagesal, "#{}".format(int(cant)), (int(cX[i]) - 10, int(cY[i])),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.4, (0, 255, 255), 1)
                        elif cant_aux==3:
                            cv2.putText(imagesal, "#{}".format(int(cant)), (int(cX[i]) - 10, int(cY[i])),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.4, (255, 255, 0), 1)
                    else:
                        cant_aux=1
                        cv2.putText(imagesal, "#{}".format(int(cant)), (int(cX[i]) - 10, int(cY[i])),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.4, (0, 0, 255), 1)
                    cant=cant+cant_aux
                    
                else:
                    cv2.putText(imagesal, "#NG", (int(cX[i]) - 10, int(cY[i])),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.4, (255, 0, 0), 1)

            if peso=="": #Verifico si se ingreso valor en cuadro de peso, si no se asigno lo pongo en 0
                peso="0"
            #Si encontro semillas verifico guardo los valoes de cantidad peso y area. Y los imprimo en la imagen
            if cant!=0:
                cv2.putText(imagesal, "Area promedio:{:.2f}mm2".format(areat/cant), (0, 60),
                            cv2.FONT_HERSHEY_SIMPLEX, 2, (0, 0, 0), 3)
            cv2.putText(imagesal, "Cantidad de semillas:{}".format(cant), (0, 120),
                        cv2.FONT_HERSHEY_SIMPLEX, 2, (0, 0, 0), 3)
            cv2.putText(imagesal, "Peso total: {:.2f}g".format(float(peso)), (0, 180),
                        cv2.FONT_HERSHEY_SIMPLEX, 2, (0, 0, 0), 3)
            pesounidad=float(peso)
            pesounidad=((pesounidad/cant)*1000)
            cv2.putText(imagesal, "Peso por semilla: {:.2f}mg".format(float(pesounidad)), (0, 240),cv2.FONT_HERSHEY_SIMPLEX, 2, (0, 0, 0), 3)
            cv2.imwrite('/home/proyecto/GranoMeter/imagensalida/salida.png',imagesal)
            #cv2.imwrite('/home/proyecto/GranoMeter/imagensalida/salidamarkers.png',markers)
            
            #Envio los valores calculados por la cola 2 a la interfaz.
            q2.put(cant)
            q2.put(areat)
            q2.put(peso)
            q2.put(pesounidad)
            input_p_ws.send('99') #Procesando 90%
            medicion=0
            
            #Exporto los valores del area de cada semilla a un excel
            wb=Workbook()
            ruta="/home/proyecto/Desktop/PlanillaAreas/SalidaAreas.xlsx"
            hoja=wb.active
            hoja.title="Areas"
            fila=1
            col_area=1
            for area in areagraf:
                hoja.cell(column=col_area, row=fila, value=area)
                fila+=1
            wb.save(filename=ruta)
            
            #Excel con Tablas
            fila2+=1
            col_area2=1
            vhora=datetime.datetime.now()
            hoja2.cell(column=col_area2, row=fila2, value=("{:02d}-{:02d}-{:04d}".format(vhora.day, vhora.month, vhora.year)))
            vhora=datetime.datetime.now()
            col_area2+=1
            hoja2.cell(column=col_area2, row=fila2, value=("{:02d}:{:02d}:{:02d}".format(vhora.hour, vhora.minute, vhora.second)))
            col_area2+=1
            hoja2.cell(column=col_area2, row=fila2, value=nombresemilla)
            col_area2+=1
            hoja2.cell(column=col_area2, row=fila2, value=muestra)
            col_area2+=1
            hoja2.cell(column=col_area2, row=fila2, value=repeticion)
            col_area2+=1
            hoja2.cell(column=col_area2, row=fila2, value=("{:.2f}".format(float(pesounidad))))
            col_area2+=1
            hoja2.cell(column=col_area2, row=fila2, value=cant)
            col_area2+=1
            hoja2.cell(column=col_area2, row=fila2, value=( "{:.2f}".format(areat/cant)))
            wb2.save(filename=ruta2)

def verimagen(): #funcion de boton ver imagen para ver la imagen con las semillas medidas
        os.system("eog --fullscreen /home/proyecto/GranoMeter/imagensalida/salida.png")

#Defino proceso de Iniciar Medicion y barra
def proceso(output_b_ba, textoprog, textopeso, textoultima,  selected, textomuestra, textorepeticion):
    global data
    q.put(1) 
    print(data)
    # Verifico si el ingreso del peso es por puerto serie o manual
    if selected.get()==1: #Ingreso por puerto serie
        q.put(data)
        q.put(textomuestra.get())
        q.put(textorepeticion.get())
        print(data)
    if selected.get()==2: #ingreso manual
        q.put(textopeso.get())
        q.put(textomuestra.get())
        q.put(textorepeticion.get())   
    q.put(ventana.combo.current())
    barra=5
    ventana.prog.set(barra) #Actualizo progreso en barra
    textoprog.set("MEDICION EN PROGRESO... Espere por favor --> {}%".format(barra))
    ventana.update_idletasks()
    barra=output_b_ba.recv()
    while(barra!='99'):
        barra=output_b_ba.recv()
        ventana.prog.set(barra) #Actualizo progreso en barra
        if(barra=='80'):
                textoultima.set(" ") #borro los datos de la ultima medicion
        textoprog.set("MEDICION EN PROGRESO... Espere por favor --> {}%".format(barra))
        ventana.update_idletasks()
    barra=0
    ventana.prog.set(barra)
    textoprog.set(" ")
    ventana.update_idletasks()
    selected.set(2)
    textopeso.config(state='normal')
    
    #tomo valores de Cola 2
    cant=q2.get()
    areat=q2.get()
    peso=q2.get()
    pesounidad=q2.get()
    
    #Muestro los valores de la ultima medicion
    textoultima.set("Ultima medicion\n\nSemilla:{}\nCantidad de semillas:{:.0f}\nArea Promedio:{:.2f}mm2\nPeso Total:{:.4f}g\nPeso por semilla:{:.2f}mg\n".format(ventana.combo.get(), cant, areat/cant, float(peso), float(pesounidad)))
    ventana.update_idletasks()
    
    #Genero nueva ventana para mostrar los valores de la medicion
    finalizado = Toplevel()
    finalizado.geometry("320x200")
    finalizado.resizable(width=False, height=False)
    finalizado.title("Medicion Finalizada")
    marco1 = ttk.Frame(finalizado, padding=(10, 10, 10, 10),relief=RAISED)
    marco1.pack(side=TOP, fill=BOTH, expand=True)
    etiq2 = Label(marco1, text="Medicion realizada", foreground='blue')
    etiq2.pack(side=TOP, padx=10)
    etiq3 = Label(marco1, text="Semilla:{}".format(ventana.combo.get()), foreground='blue')
    etiq3.place(x=50,  y=40)
    etiq4 = Label(marco1, text="Cantidad de semillas: {:.0f}".format(cant), foreground='blue')
    etiq4.place(x=50,  y=60)
    etiq5 = Label(marco1, text="Area promedio: {:.2f}mm2".format(areat/cant), foreground='blue')
    etiq5.place(x=50,  y=80)
    etiq6 = Label(marco1, text="Peso total: {:.2f}g".format(float(peso)), foreground='blue')
    etiq6.place(x=50,  y=100)
    etiq7 = Label(marco1, text="Peso por semilla: {:.2f}mg".format(float(pesounidad)), foreground='blue')
    etiq7.place(x=50,  y=120)
    botonimagen = Button(finalizado,text="VER IMAGEN",command=verimagen) #boton para ver imagen de medicion
    botonimagen.place(x=100,y=160)
    botonimagen.config(fg="white",bg="black")
    finalizado.transient(ventana)
    ventana.wait_window(finalizado)
    
# Encender Luz
def luzon():
    GPIO.output(output_luz, GPIO.HIGH)

# Apagar Luz
def luzoff():
    GPIO.output(output_luz, GPIO.LOW)

# Encender Motor1
#def motor1on():
#    GPIO.output(output_motor1, GPIO.HIGH)

# Apagar Motor1
#def motor1off():
#    GPIO.output(output_motor1, GPIO.LOW)

# Encender Motor2
#def motor2on():
#    GPIO.output(output_motor2, GPIO.HIGH)

# Apagar Motor2
#def motor2off():
#    global ventana
#    GPIO.output(output_motor2, GPIO.LOW)

# Encender ambos motores
#def motoron():
#    GPIO.output(output_motor1, GPIO.HIGH)
#    GPIO.output(output_motor2, GPIO.HIGH)

# Apagar ambos motores
#def motoroff():
#    GPIO.output(output_motor1, GPIO.LOW)
#    GPIO.output(output_motor2, GPIO.LOW)

# Salir del programa
def salir():
     global cerrartodo
     cerrartodo=1
     q.put(5) 

 #Ventana acerca de
def f_acerca():
        acerca = Toplevel()
        acerca.geometry("400x200")
        acerca.resizable(width=False, height=False)
        acerca.title("Acerca de")
        marco1 = ttk.Frame(acerca, padding=(10, 10, 10, 10),
                           relief=RAISED)
        marco1.pack(side=TOP, fill=BOTH, expand=True)
        etiq2 = Label(marco1, text="GRANOMETER Version "+__version__+"\n\nAutores:\nMatias Pariente(UTN)\nLucas Leiva(UTN)\nGerardo Zoffoli(UTN)\nIng.Agr.Dr.Daniel Miralles(FAUBA-CONICET)\n Ing.Agr.Dra.Betina Kruk(FAUBA-CONICET)", 
                      foreground='blue')
        etiq2.pack(side=TOP, padx=10)
        acerca.transient(ventana)
        ventana.wait_window(acerca) 
 
#Solo dejo ingresar valores numericos y puntos a cuadro de texto
#para ingreso de peso
def is_valid_date(action,  char,  text):
    if action !="1":
        return 1
    return char in "0123456789." and len(text) < 10

def serie(textopeso, textoprog): # si hago ingreso serie desabilito el ingreso de peso en cuadro
    global data
    textopeso.config(state='disabled')
    textoprog.set("Esperando dato serie presionar boton de balanza")
    ventana.update_idletasks()
    ser = serial.Serial('/dev/ttyUSB0',9600)
    data = ser.read(size=1)
    data = ser.read(size=7)
    data = float(data)
    #textopeso.config(state='normal')
    textoprog.set("Dato por puerto serie recibido --> {:.2f}g".format(data))
    
def manual(textopeso): # vuelvo a habilitar el ingreso
    textopeso.config(state='normal')

#interfas grafica
def interfaz(pipe_ba): 
    global cerrartodo
    global data
    output_p_ba, input_p_ba=pipe_ba
    input_p_ba.close()
    ventana.geometry('800x500')
    ventana.title("GRANOMETER")
    ventana.prog = DoubleVar() 
    
    #Defino la Etiqueta de avances
    textoprogreso=StringVar()
    textoprogreso.set("RECORDAR ENCENDER CAMARA!!")
    cajatextoprog = Label(ventana,  foreground='red')
    cajatextoprog.place(x=100, y=430)
    cajatextoprog.configure(textvariable=textoprogreso)
    
    #Defino la Etiqueta Ultima medicion
    textoultima=StringVar()
    textoultima.set(" ")
    cajatextoultima = Label(ventana, foreground='green')
    cajatextoultima.place(x=520, y=300)
    cajatextoultima.configure(textvariable=textoultima)
    
    #Defino la progress bar
    ventana.progressbar = ttk.Progressbar(ventana, mode="determinate", variable=ventana.prog)
    ventana.progressbar.place(x=100, y=450, width=350)
    
    # Agrego Etiqueta de nombre de proyecto
    ANALIZADOR=Label(ventana,text="GRANOMETER")
    ANALIZADOR.place(x=20,y=20)
    ANALIZADOR.config(font=("Verdana",30),fg="blue")
    #DEGRANOS=Label(ventana,text="DE GRANOS")
    #DEGRANOS.place(x=120,y=40)
    #DEGRANOS.config(font=("Verdana",20),fg="blue")
    
    # Agrego Etiqueta de seleccion de semilla
    SEMILLA=Label(ventana,text="Seleccionar Semilla:")
    SEMILLA.place(x=20,y=80)
    SEMILLA.config(font=("Verdana",12))
    
    # Genero Lista Desplegable con las opciones de semilla
    ventana.combo=ttk.Combobox(ventana,width=30)
    #ventana.combo['values']=("Soja","Colza","Vicia","Maiz","Trigo","Trigo Sarraceno","Poroto Blanco","Poroto Negro","Poroto Rojo","Garbanzo","Chia","Avena","Cebada","Girasol","Quinoa","Medicion Area Grande","Medicion Area Chica")
    ventana.combo['values']=("Soja","Colza","Vicia","Maiz","Trigo","Trigo Sarraceno","Poroto Blanco","Poroto Negro","Poroto Rojo","Garbanzo","Chia","Avena","Cebada","Girasol","Quinoa")
    ventana.combo.place(x=30,y=120)
    ventana.combo.set("Soja")
    
    # Agrego Etiqueta con imagen de Universidades
    imagenUTN=PhotoImage(file="/home/proyecto/GranoMeter/imagenesfondo/UTN.png")
    imagenUTN=imagenUTN.subsample(2,2)
    Label(ventana,image=imagenUTN).place(x=305,y=35)
    imagenUBA=PhotoImage(file="/home/proyecto/GranoMeter/imagenesfondo/FAUBACONICET.png")
    imagenUBA=imagenUBA.subsample(2,2)
    Label(ventana,image=imagenUBA).place(x=570,y=0)
    
    #Genero entrytext para ingresar el peso
    validatecommand=ventana.register(is_valid_date)
    textopeso=Entry(ventana, width=10, validate="key", validatecommand=(validatecommand, "%d", "%S", "%s"))
    textopeso.pack()
    textopeso.place(x=230, y=250)
    
    #Genero entradas de datos para planilla
    textomuestra=Entry(ventana, width=10)
    textomuestra.pack()
    textomuestra.place(x=50, y=320)
    MUESTRA=Label(ventana,text="Ingrese Codigo de Muestra:")
    MUESTRA.place(x=20,y=300)
    textorepeticion=Entry(ventana, width=10)
    textorepeticion.pack()
    textorepeticion.place(x=50, y=380)
    REPETICION=Label(ventana,text="Ingrese Repeticion:")
    REPETICION.place(x=20,y=360)
    
    #RadiobuttonPeso
    data=0
    selected=IntVar()
    serie_argum=partial(serie, textopeso, textoprogreso)
    manual_argum=partial(manual, textopeso)
    rad1=Radiobutton(ventana, text="Serie", value=1, variable=selected, command=serie_argum)
    rad2=Radiobutton(ventana, text="Manual", value=2, variable=selected, command=manual_argum)
    rad1.place(x=220, y=190)
    rad2.place(x=220, y=220)
    selected.set(2)
    tpeso=Label(ventana,text="Ingreso Peso (g)")
    tpeso.place(x=220,y=160)
    tpeso.config(font=("Verdana",12))
    
    # Genero etiquetas y botones de Luz
    LUZ=Label(ventana,text="LUZ")
    LUZ.place(x=90,y=180)
    LUZ.config(font=("Verdana",16))
    boton = Button(ventana,text="ENCENDER",command=luzon)
    boton.place(x=10,y=220)
    boton.config(bg="green")
    boton1 = Button(ventana,text="APAGAR",command=luzoff)
    boton1.place(x=120,y=220)
    boton1.config(fg="white",bg="red")
    
    # Genero etiquetas y botones de Motor1
    #MOTOR1=Label(ventana,text="VIBRACION 1")
    #MOTOR1.place(x=30,y=310)
    #MOTOR1.config(font=("Verdana",16))
    #boton2 = Button(ventana,text="ENCENDER",command=motor1on)
    #boton2.place(x=10,y=350)
    #boton2.config(bg="green")
    #boton3 = Button(ventana,text="APAGAR",command=motor1off)
    #boton3.place(x=120,y=350)
    #boton3.config(fg="white",bg="red")
    
    # Genero etiquetas y botones de Motor2
    #MOTOR2=Label(ventana,text="VIBRACION 2")
    #MOTOR2.place(x=300,y=310)
    #MOTOR2.config(font=("Verdana",16))
    #boton4 = Button(ventana,text="ENCENDER",command=motor2on)
    #boton4.place(x=270,y=350)
    #boton4.config(bg="green")
    #boton5 = Button(ventana,text="APAGAR",command=motor2off)
    #boton5.place(x=390,y=350)
    #boton5.config(fg="white",bg="red")
    
    # Genero etiquetas y botones de Ambos Motores
    #MOTOR2=Label(ventana,text="VIBRACION 1-2")
    #MOTOR2.place(x=550,y=310)
    #MOTOR2.config(font=("Verdana",16))
    #boton6 = Button(ventana,text="ENCENDER",command=motoron)
    #boton6.place(x=540,y=350)
    #boton6.config(bg="green")
    #boton7 = Button(ventana,text="APAGAR",command=motoroff)
    #boton7.place(x=650,y=350)
    #boton7.config(fg="white",bg="red")
    
    # Genero Boton de Iniciar Medicion
    proceso_argum=partial(proceso, output_p_ba, textoprogreso, textopeso, textoultima,  selected, textomuestra, textorepeticion)
    boton8 = Button(ventana,text="INICIAR MEDICION",command=proceso_argum)
    boton8.place(x=180,y=320)
    boton8.config(font=("Arial",24),fg="white",bg="blue")
    
    # Genero Boton de salida
    boton9 = Button(ventana,text="SALIR",command=salir)
    boton9.place(x=650,y=450)
    boton9.config(fg="white",bg="black")
    
    # Genero Barra
    barramenu = Menu(ventana)
    ventana['menu'] = barramenu
    
    # DEFINIR submenus 'Ayuda':
    menu3 = Menu(barramenu)
    barramenu.add_cascade(menu=menu3, label='Ayuda')
    menu3.add_command(label="Acerca de", command=f_acerca)
    
    while 1:
        time.sleep(0.01)
        ventana.update_idletasks()
        ventana.update()
        if cerrartodo:
            output_p_ba.close()
            ventana.destroy()
            return 0

#Programa principal
if __name__ == '__main__':
    ventana = Tk()
    #Genero colas de datos
    q= multiprocessing.Queue()
    q2= multiprocessing.Queue()
    #Genero pipe de comunicacion
    pipe_1 = multiprocessing.Pipe()
    #Genero los procesos de interfaz y watersheed , Los inicio
    p_interfaz = multiprocessing.Process(target=interfaz, args=(pipe_1, ))
    p_watersheed= multiprocessing.Process(target=watersheed, args=(range(2), q, q2,  pipe_1, ))
    p_interfaz.start()
    p_watersheed.start()
    #Al retornar cierro los procesos
    p_interfaz.join()
    p_watersheed.join()
    GPIO.cleanup() # limpio GPIO

